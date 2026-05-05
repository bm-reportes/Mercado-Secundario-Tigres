"""Procesa todos los archivos Tigres MS/TR/Listados y genera analytics.json.

Estructura:
  /Tigres/MS/AP25/<J# Rival>.xlsx       MS detalle 31 cols (vendidos confirmados)
  /Tigres/MS/CL26/<J# Rival>.xlsx       MS detalle 31 cols
  /Tigres/Listados/<J# Rival>.xlsx      Listados 8 cols (todo lo puesto en venta, status VENDIDO/EXPIRADO/Disponible)
  /Tigres/TR/<Rival>.xlsx               Mercado primario individual
  /Tigres/TR/Abonos AP26CL26.xlsx       Abonos
"""
import openpyxl
import os
import json
import re
from collections import defaultdict, Counter

ROOT = os.path.dirname(os.path.abspath(__file__))

# Factor de neto al vendedor: el archivo "Detallado" del MS expone PRECIO (col 12) = total
# bruto que paga el comprador. El reporte de "Corte" oficial de BM separa ese total en:
#   Vendedor (neto):    85.10% (lo que se transfiere al abonado)
#   Comisión Equipo:     7.46% (lo retiene el club Tigres)
#   Comisión BM:         7.44% (lo retiene Boletomóvil)
# Calculamos sobre los totales conocidos del Apertura 2025:
#   $36,523,218 / $42,920,012 = 0.85096
# Aplicamos este factor a todas las acumulaciones de dinero para reflejar netos.
VENDEDOR_FACTOR = 36523218 / 42920012   # 0.85096
EQUIPO_FACTOR = 3202446 / 42920012      # 0.07462
BM_FACTOR = 3194347 / 42920012          # 0.07444

# Modo de anonimización: si está en True, los nombres/correos/teléfonos en el output
# se enmascaran con asteriscos. Útil para compartir el reporte públicamente (ej. GitHub).
ANONYMIZE = os.environ.get('ANONYMIZE', '0') == '1'

def mask_name(s):
    if not s or not ANONYMIZE: return s
    parts = s.strip().split()
    return ' '.join(p[0] + '*' * max(len(p)-1, 2) if p else '' for p in parts)

def mask_email(s):
    if not s or not ANONYMIZE or '@' not in s: return s
    local, _, domain = s.partition('@')
    return (local[:2] + '*' * max(len(local)-2, 3)) + '@' + domain

def mask_phone(s):
    if not s or not ANONYMIZE: return s
    digits = ''.join(c for c in s if c.isdigit() or c == '+')
    if len(digits) < 6: return '***'
    return digits[:4] + '*' * (len(digits)-6) + digits[-2:]

# Mapeo rival -> torneo (para Listados/ que vienen sin prefijo de torneo)
AP25_RIVALS = {'Juarez', 'Puebla', 'America', 'Leon', 'Atlas', 'Cruz Azul', 'Necaxa', 'Tijuana', 'San Luis'}
CL26_RIVALS = {'Pumas', 'Toluca', 'Santos', 'Pachuca', 'Rayados', 'Queretaro', 'Chivas', 'Mazatlan'}

def num(v):
    if v is None or v == '-' or v == '':
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except Exception:
        return 0

def s(v):
    if v is None:
        return ''
    return str(v).strip()

def is_data_row_ms_detail(row):
    return isinstance(row[0], str) and row[0].startswith('R-')

def is_data_row_tr(row):
    return isinstance(row[0], int)

def is_data_row_listado(row):
    return isinstance(row[0], int)

def parse_jornada(filename):
    name = filename[:-5]
    m = re.match(r'(J\d+)\s+(.+)', name)
    if m:
        return m.group(1), m.group(2).strip()
    return None, name

def tournament_for_rival(rival):
    if rival in AP25_RIVALS:
        return 'AP25'
    if rival in CL26_RIVALS:
        return 'CL26'
    return 'desconocido'

# ============================================================
# 1. MS DETALLADO (vendidos confirmados con detalle vendedor/comprador)
# ============================================================
ms_detail = []

def load_ms_detail_file(path, tournament, jornada_name):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not is_data_row_ms_detail(row):
            continue
        ms_detail.append({
            'tournament': tournament,
            'jornada': jornada_name,
            'orden': s(row[0]),
            'codigo_compra': s(row[1]),
            'evento': s(row[2]),
            'evento_original': s(row[3]),
            'orden_original': s(row[4]),
            'fecha': s(row[5]),
            'tipo': s(row[6]),
            'zona': s(row[7]),
            'seccion': s(row[8]),
            'asiento': s(row[9]),
            'medio': s(row[10]),
            'metodo_pago': s(row[11]),
            'precio': num(row[12]),
            'comision': num(row[13]),
            'total': num(row[14]),
            'cb_nuevo': s(row[17]),
            'cb_restringido': s(row[18]),
            'orden_venta': s(row[19]),
            'comprador_nombre': s(row[20]),
            'comprador_correo': s(row[21]),
            'comprador_tel': s(row[22]),
            'vendedor_nombre': s(row[23]),
            'vendedor_correo': s(row[24]),
            'vendedor_tel': s(row[25]),
            'cuenta_titular': s(row[26]),
            'clabe': s(row[27]),
            'banco': s(row[28]),
            'codigo_abono': s(row[29]),
            'titular_abono': s(row[30]) if len(row) > 30 else '',
        })
        n += 1
    wb.close()
    return n

print('Cargando MS detallado por jornada...')
for tournament in ['AP25', 'CL26']:
    folder = os.path.join(ROOT, 'MS', tournament)
    if not os.path.isdir(folder):
        continue
    for f in sorted(os.listdir(folder)):
        if not f.endswith('.xlsx') or f.startswith('~'):
            continue
        _, rival = parse_jornada(f)
        n = load_ms_detail_file(os.path.join(folder, f), tournament, rival)
        print(f'  {tournament}/{f}: {n}')
print(f'Total MS detallado: {len(ms_detail)}')

# ============================================================
# 2. LISTADOS (todo lo puesto en venta — vendido, expirado, disponible)
# ============================================================
listados = []

def load_listado_file(path, tournament, rival):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not is_data_row_listado(row):
            continue
        listados.append({
            'tournament': tournament,
            'rival': rival,
            'listado_id': row[0],
            'estatus': s(row[1]).upper(),
            'zona': s(row[2]),
            'seccion': s(row[3]),
            'asiento': s(row[4]),
            'orden_venta': s(row[5]),    # RV-#####
            'orden_compra': s(row[6]),   # R-#### si vendido
            'precio': num(row[7]),
        })
        n += 1
    wb.close()
    return n

print('\nCargando Listados...')
listados_folder = os.path.join(ROOT, 'Listados')
if os.path.isdir(listados_folder):
    for f in sorted(os.listdir(listados_folder)):
        if not f.endswith('.xlsx') or f.startswith('~'):
            continue
        _, rival = parse_jornada(f)
        tournament = tournament_for_rival(rival)
        n = load_listado_file(os.path.join(listados_folder, f), tournament, rival)
        print(f'  {tournament}/{f}: {n}')
print(f'Total listados: {len(listados)}')

# ============================================================
# 3. TR — mercado primario, índice por código de barras
# ============================================================
tr_by_barcode = {}
abonos_cuenta_by_cb = {}              # cb -> cuenta del abono
abonos_cuenta_by_asiento = {}         # (zona,seccion,asiento) -> cuenta del abono
abonos_unique_cuentas = set()
abonos_unique_titulares = set()

def load_tr_file(path, source_label):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    n = 0
    is_abonos = source_label == 'Abono'
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3 or not is_data_row_tr(row):
            continue
        cb = s(row[21])
        if not cb:
            continue
        tr_by_barcode[cb] = {
            'source': source_label,
            'orden': row[0],
            'evento': s(row[2]),
            'zona': s(row[6]),
            'precio': num(row[11]),
            'cb': cb,
        }
        if is_abonos:
            cuenta = s(row[35]) if len(row) > 35 else ''
            titular = s(row[34]) if len(row) > 34 else ''
            zona = s(row[6])
            seccion = s(row[7])
            asiento = s(row[8])
            if cuenta and cuenta != '-':
                abonos_cuenta_by_cb[cb] = cuenta
                abonos_unique_cuentas.add(cuenta)
                abonos_cuenta_by_asiento[(zona, seccion, asiento)] = cuenta
            if titular and titular != '-':
                abonos_unique_titulares.add(titular)
        n += 1
    wb.close()
    return n

print('\nCargando TR mercado primario...')
for f in sorted(os.listdir(os.path.join(ROOT, 'TR'))):
    if not f.endswith('.xlsx') or f.startswith('~'):
        continue
    label = 'Abono' if 'Abonos' in f else 'Boleto Individual'
    n = load_tr_file(os.path.join(ROOT, 'TR', f), label)
    print(f'  TR/{f}: {n} ({label})')
print(f'Total TR primario: {len(tr_by_barcode)} cb · Abonos únicos: {len(abonos_unique_cuentas)} cuentas, {len(abonos_unique_titulares)} titulares')

# ============================================================
# 4. ANALYTICS
# ============================================================
analytics = {
    'meta': {
        'generated_at': '2026-05-05',
        'club': 'Tigres UANL',
        'totals_ms_orders': len(ms_detail),
        'totals_listados': len(listados),
        'totals_tr_barcodes': len(tr_by_barcode),
        'totals_abonos_cuentas': len(abonos_unique_cuentas),
        'totals_abonos_titulares': len(abonos_unique_titulares),
    }
}

# ----- 4.1 MS detalle resumen -----
def agg_ms_detail():
    total_boletos = len(ms_detail)
    total_ordenes = len(set(d['orden'] for d in ms_detail))
    total_revenue = sum((d['precio'] * VENDEDOR_FACTOR) for d in ms_detail)
    by_t = defaultdict(lambda: {'boletos': 0, 'ordenes': set(), 'revenue': 0})
    for d in ms_detail:
        t = d['tournament']
        by_t[t]['boletos'] += 1
        by_t[t]['ordenes'].add(d['orden'])
        by_t[t]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
    return {
        'boletos': total_boletos,
        'ordenes': total_ordenes,
        'revenue_total': total_revenue,
        'by_tournament': {t: {'boletos': v['boletos'], 'ordenes': len(v['ordenes']), 'revenue': v['revenue']} for t, v in by_t.items()},
    }
analytics['ms_detail_summary'] = agg_ms_detail()

# ----- 4.2 Listados (volumen total, conversion) -----
def agg_listados():
    total = len(listados)
    by_status = Counter()
    by_t = defaultdict(lambda: Counter())
    for l in listados:
        by_status[l['estatus']] += 1
        by_t[l['tournament']][l['estatus']] += 1
    vendidos = by_status.get('VENDIDO', 0)
    expirados = by_status.get('EXPIRADO', 0)
    disponibles = by_status.get('DISPONIBLE', 0)
    return {
        'total': total,
        'vendidos': vendidos,
        'expirados': expirados,
        'disponibles': disponibles,
        'conversion': vendidos / total if total else 0,
        'by_tournament': {t: dict(c) for t, c in by_t.items()},
    }
analytics['listados_summary'] = agg_listados()

# Listados por evento (jornada)
def listados_by_evento():
    agg = defaultdict(lambda: {'tournament': '', 'rival': '', 'total': 0, 'vendido': 0, 'expirado': 0, 'disponible': 0, 'revenue_listado': 0, 'revenue_vendido': 0})
    for l in listados:
        key = (l['tournament'], l['rival'])
        a = agg[key]
        a['tournament'] = l['tournament']
        a['rival'] = l['rival']
        a['total'] += 1
        a['revenue_listado'] += (l['precio'] * VENDEDOR_FACTOR)
        if l['estatus'] == 'VENDIDO':
            a['vendido'] += 1
            a['revenue_vendido'] += (l['precio'] * VENDEDOR_FACTOR)
        elif l['estatus'] == 'EXPIRADO':
            a['expirado'] += 1
        elif l['estatus'] == 'DISPONIBLE':
            a['disponible'] += 1
    rows = []
    for v in agg.values():
        v['conversion'] = v['vendido'] / v['total'] if v['total'] else 0
        rows.append(v)
    rows.sort(key=lambda x: (x['tournament'], -x['total']))
    return rows
analytics['listados_by_evento'] = listados_by_evento()

# Listados por zona — incluye conversión
def listados_by_zona():
    agg = defaultdict(lambda: {'total': 0, 'vendido': 0, 'expirado': 0, 'disponible': 0, 'revenue_vendido': 0, 'precio_total': 0})
    for l in listados:
        z = l['zona'] or '—'
        a = agg[z]
        a['total'] += 1
        a['precio_total'] += (l['precio'] * VENDEDOR_FACTOR)
        if l['estatus'] == 'VENDIDO':
            a['vendido'] += 1
            a['revenue_vendido'] += (l['precio'] * VENDEDOR_FACTOR)
        elif l['estatus'] == 'EXPIRADO':
            a['expirado'] += 1
        elif l['estatus'] == 'DISPONIBLE':
            a['disponible'] += 1
    rows = []
    for k, v in agg.items():
        rows.append({
            'zona': k,
            'total': v['total'],
            'vendido': v['vendido'],
            'expirado': v['expirado'],
            'disponible': v['disponible'],
            'conversion': v['vendido'] / v['total'] if v['total'] else 0,
            'precio_promedio': v['precio_total'] / v['total'] if v['total'] else 0,
            'revenue_vendido': v['revenue_vendido'],
        })
    rows.sort(key=lambda x: -x['total'])
    return rows
analytics['listados_by_zona'] = listados_by_zona()

# Listados por (zona × jornada) — para drill-down al hacer click en una zona
def listados_by_zona_jornada():
    agg = defaultdict(lambda: {'tournament': '', 'rival': '', 'total': 0, 'vendido': 0, 'expirado': 0, 'revenue_vendido': 0})
    for l in listados:
        key = (l['zona'], l['tournament'], l['rival'])
        a = agg[key]
        a['tournament'] = l['tournament']
        a['rival'] = l['rival']
        a['total'] += 1
        if l['estatus'] == 'VENDIDO':
            a['vendido'] += 1
            a['revenue_vendido'] += (l['precio'] * VENDEDOR_FACTOR)
        elif l['estatus'] == 'EXPIRADO':
            a['expirado'] += 1
    by_zona = defaultdict(list)
    for (zona, tor, rival), v in agg.items():
        by_zona[zona].append({
            'tournament': tor,
            'rival': rival,
            'total': v['total'],
            'vendido': v['vendido'],
            'expirado': v['expirado'],
            'conversion': v['vendido'] / v['total'] if v['total'] else 0,
            'revenue_vendido': v['revenue_vendido'],
        })
    for zona in by_zona:
        # Ordenar por torneo (AP25 primero) y luego por número de jornada (rival)
        by_zona[zona].sort(key=lambda x: (x['tournament'], x['rival']))
    return dict(by_zona)
analytics['listados_by_zona_jornada'] = listados_by_zona_jornada()

# % abonados que utilizan el MS — distinguimos LISTARON vs VENDIERON
def adopcion_abonados():
    total_cuentas = len(abonos_unique_cuentas)
    cuentas_vendieron = set()      # vendieron al menos 1 boleto en MS
    cuentas_listaron = set()       # listaron al menos 1 boleto (vendido o expirado)

    # Vendieron: por código de barras del MS detalle (boletos efectivamente vendidos)
    for d in ms_detail:
        cb = d['cb_restringido']
        if cb in abonos_cuenta_by_cb:
            cuentas_vendieron.add(abonos_cuenta_by_cb[cb])

    # Listaron: por (zona,seccion,asiento) en Listados (incluye expirados)
    for l in listados:
        key = (l['zona'], l['seccion'], l['asiento'])
        if key in abonos_cuenta_by_asiento:
            cuentas_listaron.add(abonos_cuenta_by_asiento[key])

    # Garantizar que vendieron ⊆ listaron (si vendiste, listaste — el match por cb es más preciso
    # que el de asiento, así que una cuenta que vende pero cuyo asiento no matchea por el etiquetado
    # de zonas generales debe seguir contando como "listó")
    cuentas_listaron |= cuentas_vendieron

    listaron_sin_vender = cuentas_listaron - cuentas_vendieron   # listaron pero ningún boleto se vendió
    nunca_usaron = total_cuentas - len(cuentas_listaron)
    return {
        'total_cuentas_abono': total_cuentas,
        'cuentas_vendieron_ms': len(cuentas_vendieron),       # listaron Y vendieron al menos 1
        'cuentas_listaron_ms': len(cuentas_listaron),
        'cuentas_listaron_sin_vender': len(listaron_sin_vender),
        'cuentas_nunca_usaron_ms': nunca_usaron,
        'porcentaje_listaron': len(cuentas_listaron) / total_cuentas if total_cuentas else 0,
        'porcentaje_vendieron': len(cuentas_vendieron) / total_cuentas if total_cuentas else 0,
        'porcentaje_listaron_sin_vender': len(listaron_sin_vender) / total_cuentas if total_cuentas else 0,
        'porcentaje_nunca_usaron': nunca_usaron / total_cuentas if total_cuentas else 0,
        'eficacia_listaron_a_vendieron': len(cuentas_vendieron) / len(cuentas_listaron) if cuentas_listaron else 0,
    }
analytics['adopcion_abonados'] = adopcion_abonados()

# ----- Eventos (jornadas) MS detalle -----
def by_evento_detail():
    agg = defaultdict(lambda: {'tournament': '', 'boletos': 0, 'ordenes': set(), 'revenue': 0, 'sellers': set(), 'buyers': set()})
    for d in ms_detail:
        e = d['evento'] or '—'
        if e in ('-', 'FECHA INICIO', '31/12/69'):
            continue
        agg[e]['tournament'] = d['tournament']
        agg[e]['boletos'] += 1
        agg[e]['ordenes'].add(d['orden'])
        agg[e]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
        if d['vendedor_correo']:
            agg[e]['sellers'].add(d['vendedor_correo'].lower())
        if d['comprador_correo']:
            agg[e]['buyers'].add(d['comprador_correo'].lower())
    rows = []
    for k, v in agg.items():
        rows.append({
            'evento': k,
            'tournament': v['tournament'],
            'boletos': v['boletos'],
            'ordenes': len(v['ordenes']),
            'revenue': v['revenue'],
            'sellers_unique': len(v['sellers']),
            'buyers_unique': len(v['buyers']),
        })
    def sk(r):
        m = re.search(r'J-?(\d+)', r['evento'])
        return (r['tournament'], int(m.group(1)) if m else 0)
    rows.sort(key=sk)
    return rows
analytics['by_evento_detail'] = by_evento_detail()

def top_sellers(limit=100):
    agg = defaultdict(lambda: {'nombre': '', 'correo': '', 'tels': Counter(), 'boletos': 0, 'ordenes': set(), 'revenue': 0, 'eventos': set(), 'abonos': set()})
    for d in ms_detail:
        key = d['vendedor_correo'].lower() or d['vendedor_nombre'].lower()
        if not key:
            continue
        a = agg[key]
        a['nombre'] = d['vendedor_nombre']
        a['correo'] = d['vendedor_correo']
        if d['vendedor_tel']:
            a['tels'][d['vendedor_tel']] += 1
        a['boletos'] += 1
        a['ordenes'].add(d['orden'])
        a['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
        a['eventos'].add(d['evento'])
        if d['cb_restringido']:
            a['abonos'].add(d['cb_restringido'])
    rows = []
    for v in agg.values():
        tel = v['tels'].most_common(1)[0][0] if v['tels'] else ''
        rows.append({
            'nombre': mask_name(v['nombre']),
            'correo': mask_email(v['correo']),
            'telefono': mask_phone(tel),
            'boletos': v['boletos'],
            'ordenes': len(v['ordenes']),
            'abonos': len(v['abonos']),
            'eventos': len(v['eventos']),
            'revenue': v['revenue'],
        })
    rows.sort(key=lambda x: -x['boletos'])
    return rows[:limit]
analytics['top_sellers'] = top_sellers(100)

def top_buyers(limit=100):
    agg = defaultdict(lambda: {'nombre': '', 'correo': '', 'tels': Counter(), 'boletos': 0, 'ordenes': set(), 'revenue': 0, 'eventos': set()})
    for d in ms_detail:
        key = d['comprador_correo'].lower() or d['comprador_nombre'].lower()
        if not key:
            continue
        a = agg[key]
        a['nombre'] = d['comprador_nombre']
        a['correo'] = d['comprador_correo']
        if d['comprador_tel']:
            a['tels'][d['comprador_tel']] += 1
        a['boletos'] += 1
        a['ordenes'].add(d['orden'])
        a['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
        a['eventos'].add(d['evento'])
    rows = []
    for v in agg.values():
        tel = v['tels'].most_common(1)[0][0] if v['tels'] else ''
        rows.append({
            'nombre': mask_name(v['nombre']),
            'correo': mask_email(v['correo']),
            'telefono': mask_phone(tel),
            'boletos': v['boletos'],
            'ordenes': len(v['ordenes']),
            'revenue': v['revenue'],
            'eventos': len(v['eventos']),
        })
    rows.sort(key=lambda x: -x['boletos'])
    return rows[:limit]
analytics['top_buyers'] = top_buyers(100)

def by_zona_ms():
    agg = defaultdict(lambda: {'orders': 0, 'revenue': 0, 'precio_acum': 0})
    for d in ms_detail:
        z = d['zona'] or '—'
        agg[z]['orders'] += 1
        agg[z]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
        agg[z]['precio_acum'] += (d['precio'] * VENDEDOR_FACTOR)
    rows = [{'zona': k, 'orders': v['orders'], 'revenue': v['revenue'], 'precio_avg': v['precio_acum']/v['orders'] if v['orders'] else 0} for k, v in agg.items()]
    rows.sort(key=lambda x: -x['orders'])
    return rows
analytics['by_zona'] = by_zona_ms()

def by_medio():
    agg = defaultdict(lambda: {'orders': 0, 'revenue': 0})
    for d in ms_detail:
        m = d['medio'] or '—'
        agg[m]['orders'] += 1
        agg[m]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
    rows = [{'medio': k, **v} for k, v in agg.items()]
    rows.sort(key=lambda x: -x['orders'])
    return rows
analytics['by_medio'] = by_medio()

def by_metodo_pago():
    agg = defaultdict(lambda: {'orders': 0, 'revenue': 0})
    for d in ms_detail:
        m = d['metodo_pago'] or '—'
        agg[m]['orders'] += 1
        agg[m]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
    rows = [{'metodo': k, **v} for k, v in agg.items()]
    rows.sort(key=lambda x: -x['orders'])
    return rows
analytics['by_metodo_pago'] = by_metodo_pago()

def by_banco():
    agg = defaultdict(lambda: {'orders': 0, 'revenue': 0, 'sellers': set()})
    for d in ms_detail:
        b = d['banco'] or '—'
        agg[b]['orders'] += 1
        agg[b]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
        agg[b]['sellers'].add(d['vendedor_correo'].lower())
    rows = [{'banco': k, 'orders': v['orders'], 'revenue': v['revenue'], 'sellers': len(v['sellers'])} for k, v in agg.items()]
    rows.sort(key=lambda x: -x['orders'])
    return rows
analytics['by_banco'] = by_banco()

def by_tipo():
    agg = defaultdict(lambda: {'orders': 0, 'revenue': 0})
    for d in ms_detail:
        t = d['tipo'] or '—'
        agg[t]['orders'] += 1
        agg[t]['revenue'] += (d['precio'] * VENDEDOR_FACTOR)
    rows = [{'tipo': k, **v} for k, v in agg.items()]
    rows.sort(key=lambda x: -x['orders'])
    return rows
analytics['by_tipo'] = by_tipo()

def match_tr():
    matched = 0
    no_match = 0
    by_origin = defaultdict(int)
    by_origin_revenue = defaultdict(float)
    for d in ms_detail:
        cb = d['cb_restringido']
        if cb in tr_by_barcode:
            matched += 1
            by_origin[tr_by_barcode[cb]['source']] += 1
            by_origin_revenue[tr_by_barcode[cb]['source']] += (d['precio'] * VENDEDOR_FACTOR)
        else:
            no_match += 1
            by_origin['Sin match'] += 1
    return {
        'matched': matched,
        'no_match': no_match,
        'by_origin': dict(by_origin),
        'by_origin_revenue': dict(by_origin_revenue),
    }
analytics['tr_match'] = match_tr()

def concentracion_vendedores():
    boletos_counter = Counter()
    ordenes_set = defaultdict(set)
    for d in ms_detail:
        k = d['vendedor_correo'].lower() or d['vendedor_nombre'].lower()
        if k:
            boletos_counter[k] += 1
            ordenes_set[k].add(d['orden'])
    total_boletos = sum(boletos_counter.values())
    sorted_b = sorted(boletos_counter.values(), reverse=True)
    sorted_o = sorted([len(s) for s in ordenes_set.values()], reverse=True)
    n = len(sorted_b)
    top1pct = max(1, n // 100)
    top10pct = max(1, n // 10)
    return {
        'unique_sellers': n,
        'total_boletos': total_boletos,
        'avg_boletos_per_seller': total_boletos / n if n else 0,
        'avg_ordenes_per_seller': sum(sorted_o) / n if n else 0,
        'top_1pct_share': sum(sorted_b[:top1pct]) / total_boletos if total_boletos else 0,
        'top_10pct_share': sum(sorted_b[:top10pct]) / total_boletos if total_boletos else 0,
        'sellers_with_1_boleto': sum(1 for c in sorted_b if c == 1),
        'sellers_with_2plus': sum(1 for c in sorted_b if c >= 2),
        'sellers_with_5plus': sum(1 for c in sorted_b if c >= 5),
        'sellers_with_10plus': sum(1 for c in sorted_b if c >= 10),
        'sellers_with_20plus': sum(1 for c in sorted_b if c >= 20),
        'max_boletos_single_seller': max(sorted_b) if sorted_b else 0,
        'max_ordenes_single_seller': max(sorted_o) if sorted_o else 0,
    }
analytics['concentracion_vendedores'] = concentracion_vendedores()


# Save
out_path = os.path.join(ROOT, 'analytics.json')
with open(out_path, 'w', encoding='utf-8') as f:
    json.dump(analytics, f, ensure_ascii=False, default=str)
print(f'\nOK -> {out_path}  ({os.path.getsize(out_path):,} bytes)')
